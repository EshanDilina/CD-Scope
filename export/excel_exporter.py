"""
cd_scope.export.excel_exporter
────────────────────────────────
Generate a formatted .xlsx workbook from CD_SCOPE data.

Sheets
  Summary       — KPI table (Cp/Cpk, PASS/FAIL)
  CD_Data       — per-site/per-image CD table with auto-filter
  LWR_Data      — LWR/LER/PSD table + per-row edge positions
  Dose_Focus    — scanner exposure data with formulas
  Batch_Results — multi-image analysis table
  Charts        — all charts placed in a grid (no human steps)
"""
from __future__ import annotations
import datetime
from pathlib import Path

import numpy as np
from scipy import stats

from cd_scope.constants import TARGET_CD, USL, LSL, EXCEL_OK
_COLORS = dict(
    DARK="FF0F1520", MID="FF141C2A", CYAN="FF00D4FF", GREEN="FF00FF88",
    AMBER="FFFFB300", RED="FFFF3355", WHITE="FFE8F4FF", DIM="FF7A9AB8",
    PURPLE="FF9966FF",
)


def _C(name): return _COLORS[name]


if EXCEL_OK:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, LineChart, ScatterChart, Reference, Series
    from openpyxl.utils import get_column_letter


# ── Cell helpers ──────────────────────────────────────────────────────────────

def _font(bold=False, color=_C('WHITE'), size=10, name="Arial"):
    return Font(name=name, bold=bold, color=color, size=size)

def _fill(hex_color: str):
    return PatternFill("solid", fgColor=hex_color)

def _border():
    s = Side(style='thin', color=_C('DARK'))
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _hdr(ws, row, col, text, width=None, bg=_C('MID'), fg=_C('CYAN')):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = _font(bold=True, color=fg, size=10)
    c.fill      = _fill(bg)
    c.alignment = _align()
    c.border    = _border()
    if width:
        letter = get_column_letter(col)
        if ws.column_dimensions[letter].width < width:
            ws.column_dimensions[letter].width = width
    return c

def _cell(ws, row, col, value, color=_C('WHITE'), bold=False,
          fmt=None, bg=_C('DARK')):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = _font(bold=bold, color=color)
    c.fill      = _fill(bg)
    c.alignment = _align()
    c.border    = _border()
    if fmt:
        c.number_format = fmt
    return c


# ── Main exporter ─────────────────────────────────────────────────────────────

class MetroscanExcelExporter:
    """
    Build a complete Excel workbook from CD_SCOPE measurement data.

    Attributes can be set before calling .export().
    """

    def __init__(self):
        self.sites:           list = []
        self.batch_records:   list = []
        self.scanner_fields:  list = []
        self.edge_result           = None
        self.recipe_name:     str  = ""
        self.lot_id:          str  = ""
        self.wafer_id:        str  = ""
        # Sheet toggles
        self.include_summary     = True
        self.include_cd_data     = True
        self.include_lwr_data    = True
        self.include_dose_focus  = True
        self.include_batch       = True
        self.include_charts      = True

    def export(self, path: str) -> str:
        if not EXCEL_OK:
            raise RuntimeError("Install openpyxl: pip install openpyxl")
        wb = Workbook()
        wb.remove(wb.active)
        refs: dict[str, object] = {}

        if self.include_summary:
            ws = wb.create_sheet("Summary")
            self._write_summary(ws)
            refs['Summary'] = ws

        if self.include_cd_data and (self.sites or self.batch_records):
            ws = wb.create_sheet("CD_Data")
            self._write_cd_data(ws)
            refs['CD_Data'] = ws

        if self.include_lwr_data and self.edge_result:
            ws = wb.create_sheet("LWR_Data")
            self._write_lwr_data(ws)

        if self.include_dose_focus and self.scanner_fields:
            ws = wb.create_sheet("Dose_Focus")
            self._write_dose_focus(ws)

        if self.include_batch and self.batch_records:
            ws = wb.create_sheet("Batch_Results")
            self._write_batch(ws)

        if self.include_charts:
            ws = wb.create_sheet("Charts")
            self._write_charts(ws, wb, refs)

        wb.save(path)
        return path

    # ── Summary ────────────────────────────────────────────────────────────────

    def _write_summary(self, ws):
        ws.sheet_view.showGridLines = False
        for col, w in [(1, 28), (2, 16), (3, 16), (4, 16)]:
            ws.column_dimensions[get_column_letter(col)].width = w

        row = 1
        c = ws.cell(row=row, column=1, value="CD_SCOPE v1.0 — Analysis Summary")
        c.font = Font(name="Arial", bold=True, color=_C('CYAN'), size=14)
        c.fill = _fill("FF0B0F17")
        ws.merge_cells(f"A{row}:D{row}")
        c.alignment = _align('left')
        row += 1

        for k, v in [("Recipe", self.recipe_name or "—"),
                     ("Lot ID", self.lot_id or "—"),
                     ("Wafer ID", self.wafer_id or "—"),
                     ("Date", datetime.datetime.now().strftime("%Y-%m-%d %H:%M"))]:
            _cell(ws, row, 1, k, _C('DIM'), bg="FF0B0F17")
            c = ws.cell(row=row, column=2, value=v)
            c.font  = _font(color=_C('WHITE')); c.fill = _fill("FF0B0F17")
            ws.merge_cells(f"B{row}:D{row}")
            row += 1

        row += 1
        for col, h in enumerate(["METRIC", "VALUE", "SPEC", "STATUS"], 1):
            _hdr(ws, row, col, h)
        row += 1

        for m in self._collect_metrics():
            bg  = ("FF071507" if m['status']=='PASS'
                   else "FF1A0505" if m['status']=='FAIL' else _C('DARK'))
            sc  = (_C('GREEN') if m['status']=='PASS'
                   else _C('RED') if m['status']=='FAIL' else _C('WHITE'))
            _cell(ws, row, 1, m['name'],   _C('DIM'),   bg=bg)
            _cell(ws, row, 2, m['value'],  _C('WHITE'),
                  bold=True, bg=bg,
                  fmt='0.000' if isinstance(m['value'], float) else None)
            _cell(ws, row, 3, m['spec'],   _C('DIM'),   bg=bg)
            _cell(ws, row, 4, m['status'], sc, bold=True, bg=bg)
            row += 1

        ws.freeze_panes = f"A{row+1}"

    def _collect_metrics(self) -> list[dict]:
        metrics = []
        cds = ([s.cd_mean for s in self.sites if s.cd_mean > 0] +
               [r.cd_mean for r in self.batch_records if r.cd_mean > 0])
        lwrs = ([s.lwr for s in self.sites if s.lwr > 0] +
                [r.lwr_3s for r in self.batch_records if r.lwr_3s > 0])
        if cds:
            mu = np.mean(cds); sg = np.std(cds, ddof=1) if len(cds)>1 else 0
            cp  = (USL-LSL)/(6*sg) if sg>0 else 999
            cpu = (USL-mu)/(3*sg)  if sg>0 else 999
            cpl = (mu-LSL)/(3*sg)  if sg>0 else 999
            cpk = min(cpu, cpl)
            metrics += [
                {'name':'CD Mean (nm)', 'value':round(mu,3), 'spec':'30–34 nm',
                 'status':'PASS' if 30<=mu<=34 else 'FAIL'},
                {'name':'CD 3σ (nm)',   'value':round(3*sg,3),'spec':'<6 nm',
                 'status':'PASS' if 3*sg<6 else 'FAIL'},
                {'name':'Cp',           'value':round(cp,3),  'spec':'≥1.33',
                 'status':'PASS' if cp>=1.33 else 'FAIL'},
                {'name':'Cpk',          'value':round(cpk,3), 'spec':'≥1.33',
                 'status':'PASS' if cpk>=1.33 else 'FAIL'},
                {'name':'N Sites',      'value':len(cds), 'spec':'—', 'status':'—'},
            ]
        if lwrs:
            m = np.mean(lwrs)
            metrics.append({'name':'LWR Mean (nm)', 'value':round(m,3),
                             'spec':'<4 nm', 'status':'PASS' if m<4 else 'FAIL'})
        return metrics

    # ── CD Data ────────────────────────────────────────────────────────────────

    def _write_cd_data(self, ws):
        ws.sheet_view.showGridLines = False
        headers = ["#","SITE","X_mm","Y_mm","CD_MEAN","CD_STD",
                   "LWR","PITCH","SPACE","DUTY%","STATUS"]
        widths  = [5,10,10,10,12,12,10,10,10,10,10]
        row = 1
        _hdr(ws, row, 1, "CD Measurement Data", 28, bg="FF0B0F17")
        ws.merge_cells(f"A{row}:{get_column_letter(len(headers))}{row}")
        row += 1
        for col,(h,w) in enumerate(zip(headers, widths), 1):
            _hdr(ws, row, col, h, w)
        row += 1; data_start = row

        for i, s in enumerate(self.sites):
            duty = s.cd_mean / s.pitch * 100 if s.pitch > 0 else 0
            bg   = _C('DARK') if i%2==0 else "FF111825"
            fc   = _C('GREEN') if s.status=='PASS' else _C('RED')
            for col, v in enumerate([i+1, s.site_id, s.x_mm, s.y_mm,
                                      s.cd_mean, s.cd_std, s.lwr,
                                      s.pitch, s.space, round(duty,1), s.status], 1):
                cc = fc if col == 11 else _C('WHITE')
                _cell(ws, row, col, v, cc, col==11,
                      '0.000' if col in (5,6,7,8,9) else None, bg)
            row += 1

        for i, rec in enumerate(self.batch_records):
            bg = _C('DARK') if i%2==0 else "FF111825"
            fc = _C('GREEN') if rec.status=='PASS' else (
                 _C('RED') if rec.status=='FAIL' else _C('AMBER'))
            for col, v in enumerate([i+1, rec.site_id, rec.x_mm, rec.y_mm,
                                      rec.cd_mean, rec.cd_std, rec.lwr_3s,
                                      rec.pitch_mean, rec.space_mean, 0, rec.status], 1):
                cc = fc if col == 11 else _C('WHITE')
                _cell(ws, row, col, v, cc, col==11,
                      '0.000' if col in (5,6,7,8,9) else None, bg)
            row += 1

        # Summary formulas
        row += 1
        for lbl, formula in [("Mean", f"=AVERAGE(E{data_start}:E{row-2})"),
                              ("Std",  f"=STDEV(E{data_start}:E{row-2})"),
                              ("Min",  f"=MIN(E{data_start}:E{row-2})"),
                              ("Max",  f"=MAX(E{data_start}:E{row-2})")]:
            _cell(ws, row, 4, lbl, _C('CYAN'), True, bg=_C('MID'))
            _cell(ws, row, 5, formula, _C('WHITE'), bg=_C('MID'), fmt='0.000')
            row += 1

        ws.freeze_panes = "A4"
        ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{data_start-1}"

    # ── LWR Data ───────────────────────────────────────────────────────────────

    def _write_lwr_data(self, ws):
        ws.sheet_view.showGridLines = False
        r = self.edge_result
        if not r: return
        row = 1
        _hdr(ws, row, 1, "LWR / LER Analysis", 28, bg="FF0B0F17", fg=_C('PURPLE'))
        ws.merge_cells(f"A{row}:F{row}"); row += 2

        for k, v, spec in [
            ("LWR 3σ (nm)", r.lwr_3s, "< 4.0 nm"),
            ("LER Left 3σ", r.ler_l_3s, "< 3.0 nm"),
            ("LER Right 3σ", r.ler_r_3s, "< 3.0 nm"),
            ("Hurst exponent", r.hurst, "0.5–0.8"),
            ("Corr. length (nm)", r.corr_len, "nm"),
            ("Edge slope L (°)", r.edge_slope_l, "°"),
            ("Edge slope R (°)", r.edge_slope_r, "°"),
            ("Algorithm", r.algo, ""),
        ]:
            _cell(ws, row, 1, k, _C('DIM'), bg=_C('MID'))
            _cell(ws, row, 2, v, _C('WHITE'), bold=True,
                  fmt='0.000' if isinstance(v, float) else None, bg=_C('MID'))
            _cell(ws, row, 3, spec, _C('DIM'), bg=_C('MID'))
            row += 1

        # PSD table
        if len(r.psd_freq) > 0:
            row += 1
            _hdr(ws, row, 1, "Freq (µm⁻¹)", 18)
            _hdr(ws, row, 2, "PSD (nm²·µm)", 18)
            row += 1
            for freq, pwr in zip(r.psd_freq, r.psd_power):
                _cell(ws, row, 1, round(float(freq), 4), bg=_C('DARK'), fmt='0.0000')
                _cell(ws, row, 2, round(float(pwr),  6), bg=_C('DARK'), fmt='0.000000')
                row += 1

    # ── Dose/Focus ─────────────────────────────────────────────────────────────

    def _write_dose_focus(self, ws):
        ws.sheet_view.showGridLines = False
        if not self.scanner_fields: return
        row = 1
        _hdr(ws, row, 1, "Scanner Dose / Focus Data", 28, bg="FF0B0F17", fg=_C('AMBER'))
        ws.merge_cells(f"A{row}:J{row}"); row += 1
        headers = ["FIELD","X_mm","Y_mm","DOSE","FOCUS","NA","SIGMA","λ","LOT","WAFER"]
        widths  = [10,10,10,16,14,8,8,8,12,10]
        for col,(h,w) in enumerate(zip(headers, widths), 1):
            _hdr(ws, row, col, h, w)
        row += 1; data_start = row
        for i, fe in enumerate(self.scanner_fields):
            bg = _C('DARK') if i%2==0 else "FF111825"
            for col, v in enumerate([
                fe.field_id, fe.x_mm, fe.y_mm, fe.dose, fe.focus,
                fe.na or "—", fe.sigma or "—", fe.wavelength,
                fe.lot_id, fe.wafer_id
            ], 1):
                _cell(ws, row, col, v, _C('WHITE'),
                      fmt='0.000' if col in (2,3,4,5) else None, bg=bg)
            row += 1
        ws.freeze_panes = "A4"

    # ── Batch Results ──────────────────────────────────────────────────────────

    def _write_batch(self, ws):
        ws.sheet_view.showGridLines = False
        if not self.batch_records: return
        row = 1
        _hdr(ws, row, 1, "Batch Analysis Results", 28, bg="FF0B0F17")
        ws.merge_cells(f"A{row}:M{row}"); row += 1
        headers = ["#","SITE","IMAGE","PATTERN","X","Y","DOSE","FOCUS",
                   "CD","CD_σ","LWR","PITCH","STATUS"]
        widths  = [5,8,24,16,8,8,8,8,10,10,8,8,8]
        for col,(h,w) in enumerate(zip(headers, widths), 1):
            _hdr(ws, row, col, h, w)
        row += 1
        for i, rec in enumerate(self.batch_records):
            bg = _C('DARK') if i%2==0 else "FF111825"
            fc = (_C('GREEN') if rec.status=='PASS'
                  else _C('RED') if rec.status=='FAIL' else _C('AMBER'))
            img_name = Path(rec.image_path).name if rec.image_path else "—"
            for col, v in enumerate([
                i+1, rec.site_id, img_name, rec.pattern_type,
                rec.x_mm, rec.y_mm, rec.dose, rec.focus,
                rec.cd_mean, rec.cd_std, rec.lwr_3s, rec.pitch_mean, rec.status
            ], 1):
                cc = fc if col == 13 else _C('WHITE')
                _cell(ws, row, col, v, cc, col==13,
                      '0.000' if col in (5,6,7,8,9,10,11,12) else None, bg)
            row += 1
        ws.freeze_panes = "A4"
        ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{row-1}"

    # ── Charts ─────────────────────────────────────────────────────────────────

    def _write_charts(self, ws, wb, refs):
        ws.sheet_view.showGridLines = False
        c = ws.cell(row=1, column=1, value="CD_SCOPE Charts")
        c.font = Font(name="Arial", bold=True, color=_C('CYAN'), size=14)
        c.fill = _fill("FF0B0F17")
        ws.merge_cells("A1:Z1")
        charts: list[tuple] = []  # (chart, anchor)

        # CD Trend
        n = len(self.sites) + len(self.batch_records)
        if n > 0 and 'CD_Data' in refs:
            ws_cd = refs['CD_Data']
            chart = LineChart()
            chart.title = "CD Measurement Trend"
            chart.style = 2
            chart.y_axis.title = "CD (nm)"
            chart.x_axis.title = "Site #"
            chart.height = 12; chart.width = 20
            data_ref = Reference(ws_cd, min_col=5, min_row=3, max_row=3+n)
            chart.add_data(data_ref, titles_from_data=True)
            # Target column
            for ri in range(4, 4+n):
                ws_cd.cell(row=ri, column=14, value=TARGET_CD)
            ws_cd.cell(row=3, column=14, value="Target")
            chart.add_data(Reference(ws_cd, min_col=14, min_row=3, max_row=3+n),
                           titles_from_data=True)
            chart.series[0].graphicalProperties.line.solidFill = "FF00D4FF"
            if len(chart.series) > 1:
                chart.series[1].graphicalProperties.line.solidFill = "FFFFB300"
            charts.append((chart, "B3"))

        # CD Histogram (data written to hidden cols in Charts sheet)
        cds = ([s.cd_mean for s in self.sites if s.cd_mean>0] +
               [r.cd_mean for r in self.batch_records if r.cd_mean>0])
        if cds:
            bins = np.linspace(min(cds)-1, max(cds)+1, 15)
            counts, _ = np.histogram(cds, bins=bins)
            sr = 40
            ws.cell(row=sr, column=25, value="CD_bin")
            ws.cell(row=sr, column=26, value="Count")
            for j, (ctr, cnt) in enumerate(zip((bins[:-1]+bins[1:])/2, counts)):
                ws.cell(row=sr+1+j, column=25, value=round(float(ctr), 2))
                ws.cell(row=sr+1+j, column=26, value=int(cnt))
            chart2 = BarChart()
            chart2.title = "CD Distribution"
            chart2.type  = "col"
            chart2.style = 2
            chart2.y_axis.title = "Count"
            chart2.x_axis.title = "CD (nm)"
            chart2.height = 12; chart2.width = 20
            chart2.add_data(Reference(ws, min_col=26, min_row=sr, max_row=sr+len(counts)),
                            titles_from_data=True)
            chart2.set_categories(Reference(ws, min_col=25, min_row=sr+1, max_row=sr+len(counts)))
            chart2.series[0].graphicalProperties.solidFill = "FF0B6880"
            charts.append((chart2, "M3"))

        # LWR bar
        if self.sites:
            n2 = len(self.sites)
            chart3 = BarChart()
            chart3.title = "LWR 3σ per Site"
            chart3.type  = "col"
            chart3.style = 2
            chart3.y_axis.title = "LWR 3σ (nm)"
            chart3.height = 12; chart3.width = 20
            if 'CD_Data' in refs:
                chart3.add_data(Reference(refs['CD_Data'], min_col=7,
                                          min_row=3, max_row=3+n2),
                                titles_from_data=True)
                chart3.series[0].graphicalProperties.solidFill = "FF9966FF"
            charts.append((chart3, "B22"))

        # Dose vs CD scatter
        d_pts = [(r.dose, r.cd_mean) for r in self.batch_records
                 if r.dose > 0 and r.cd_mean > 0]
        if d_pts:
            dc, dcd = zip(*d_pts)
            sr2 = 60
            ws.cell(row=sr2, column=28, value="Dose")
            ws.cell(row=sr2, column=29, value="CD")
            for j, (d, c) in enumerate(zip(dc, dcd)):
                ws.cell(row=sr2+1+j, column=28, value=round(d, 3))
                ws.cell(row=sr2+1+j, column=29, value=round(c, 3))
            chart4 = ScatterChart()
            chart4.title = "Dose vs CD"
            chart4.style = 2
            chart4.y_axis.title = "CD (nm)"
            chart4.x_axis.title = "Dose (mJ/cm²)"
            chart4.height = 12; chart4.width = 20
            s4 = Series(
                Reference(ws, min_col=29, min_row=sr2, max_row=sr2+len(d_pts)),
                Reference(ws, min_col=28, min_row=sr2+1, max_row=sr2+len(d_pts)),
                title="CD vs Dose")
            s4.marker.symbol = "circle"
            s4.marker.graphicalProperties.solidFill = "FF00FF88"
            s4.graphicalProperties.line.noFill = True
            chart4.series.append(s4)
            charts.append((chart4, "M22"))

        # PSD chart
        r = self.edge_result
        if r and len(r.psd_freq) > 2:
            nz = r.psd_freq > 0
            fq, pw = r.psd_freq[nz], r.psd_power[nz]
            sr3 = 80
            ws.cell(row=sr3, column=32, value="Freq")
            ws.cell(row=sr3, column=33, value="PSD")
            for j, (f, p) in enumerate(zip(fq, pw)):
                ws.cell(row=sr3+1+j, column=32, value=round(float(f), 4))
                ws.cell(row=sr3+1+j, column=33, value=round(float(p), 6))
            chart5 = LineChart()
            chart5.title = "LWR Power Spectral Density"
            chart5.style = 2
            chart5.y_axis.title = "PSD (nm²·µm)"
            chart5.x_axis.title = "Spatial Freq (µm⁻¹)"
            chart5.height = 12; chart5.width = 20
            chart5.add_data(Reference(ws, min_col=33, min_row=sr3, max_row=sr3+len(fq)),
                            titles_from_data=True)
            chart5.series[0].graphicalProperties.line.solidFill = "FF9966FF"
            charts.append((chart5, "B41"))

        for chart, anchor in charts:
            ws.add_chart(chart, anchor)
